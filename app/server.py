import os
import uuid
import shutil
from datetime import datetime
from typing import List, Optional
from fastapi import FastAPI, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from dotenv import load_dotenv

load_dotenv()

from app.database import get_db, init_db
from app.models import (Order, Factory, OrderStatus, User, ActivityLog,
                        SiteSettings, PriceBatch, PriceItem, OrderAttachment, OrderHistory)
from app.auth import verify_password, hash_password, make_token, decode_token
from app.email_service import send_factory_email, send_admin_notification
from app.price_parser import parse_price_file

SITE_URL = os.getenv("SITE_URL", "http://localhost:8000")
COMMISSION_RATE = 0.03


def send_telegram_message(chat_id, text: str):
    """Отправить сообщение менеджеру в Telegram через Bot API."""
    import urllib.request, json as _json
    token = os.getenv("BOT_TOKEN", "")
    if not token or not chat_id:
        return
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = _json.dumps({"chat_id": str(chat_id), "text": text, "parse_mode": "Markdown"}).encode()
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
    try:
        urllib.request.urlopen(req, timeout=5)
    except Exception:
        pass

app = FastAPI(title="Мебельные заявки")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
templates.env.filters["dt"]   = lambda x: x.strftime("%d.%m.%Y %H:%M") if x else "—"
templates.env.filters["date"] = lambda x: x.strftime("%d.%m.%Y") if x else "—"
templates.env.filters["rub"]  = lambda x: f"{x:,.0f} ₽".replace(",", " ") if x else "—"

STATIC_DIR  = os.path.join(BASE_DIR, "static")
UPLOADS_DIR = os.environ.get("UPLOADS_DIR", "/app/data/uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

STATUS_LABELS = {
    "new": "Новая",
    "sent_to_factory": "Отправлено",
    "accepted": "Принято",
    "shipped": "Отгружено",
    "delivered": "Доставлено",
    "rejected": "Отклонено",
    "cancelled": "Отменена",
}

PAYMENT_METHODS = ["Наличные", "Безналичный расчёт", "Перевод", "Рассрочка", "Кредит"]


# ── helpers ───────────────────────────────────────────────────────────────────

def get_user(request: Request, db: Session):
    token = request.cookies.get("session")
    if not token:
        return None
    data = decode_token(token)
    if not data:
        return None
    user = db.query(User).filter(User.id == data["user_id"]).first()
    return user if user and user.is_active else None


def get_settings(db: Session):
    s = db.query(SiteSettings).first()
    if not s:
        s = SiteSettings(); db.add(s); db.commit()
    return s


def ctx(request: Request, db: Session, user: User, **extra):
    import time
    from datetime import date
    s = get_settings(db)
    trash_q = db.query(Order).filter(Order.deleted_at != None)
    if user.role != "admin":
        trash_q = trash_q.filter(Order.manager_username == user.username)
    trash_count = trash_q.count()
    return {"request": request, "me": user, "theme": s.theme,
            "primary_color": s.primary_color, "status_labels": STATUS_LABELS,
            "trash_count": trash_count, "now_ts": time.time(),
            "today_str": date.today().isoformat(), **extra}


def log(db: Session, user: User, action: str, request: Request = None):
    ip = request.client.host if request else None
    db.add(ActivityLog(username=user.username, role=user.role,
                       action=action, ip_address=ip))
    db.commit()


def log_history(db: Session, order_id: int, username: str,
                field: str = None, old_value=None, new_value=None, comment: str = None):
    db.add(OrderHistory(order_id=order_id, username=username, field=field,
                        old_value=str(old_value) if old_value is not None else None,
                        new_value=str(new_value) if new_value is not None else None,
                        comment=comment))


def _parse_float(val: str):
    if not val:
        return None
    try:
        return float(str(val).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


@app.on_event("startup")
def startup():
    init_db()
    # Автоочистка корзины: физически удаляем заявки старше 90 дней
    from datetime import timedelta
    db = next(get_db())
    try:
        cutoff = datetime.utcnow() - timedelta(days=90)
        expired = db.query(Order).filter(Order.deleted_at != None, Order.deleted_at < cutoff).all()
        for order in expired:
            order_dir = os.path.join(UPLOADS_DIR, str(order.id))
            if os.path.exists(order_dir):
                shutil.rmtree(order_dir)
            db.query(OrderAttachment).filter(OrderAttachment.order_id == order.id).delete()
            db.delete(order)
        if expired:
            db.commit()
    finally:
        db.close()


# ── auth ──────────────────────────────────────────────────────────────────────

@app.get("/")
def root(request: Request):
    return RedirectResponse("/orders" if request.cookies.get("session") else "/login")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...),
          db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    if not user or not user.is_active or not verify_password(password, user.password_hash):
        return templates.TemplateResponse("login.html",
            {"request": request, "error": "Неверный логин или пароль"})
    user.last_login_at = datetime.utcnow()
    db.commit()
    log(db, user, "Вход в систему", request)
    resp = RedirectResponse("/orders", status_code=303)
    resp.set_cookie("session", make_token(user.id, user.username, user.role),
                    httponly=True, max_age=86400 * 7)
    return resp


@app.get("/logout")
def logout(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if user:
        log(db, user, "Выход из системы", request)
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie("session")
    return resp


# ── orders ────────────────────────────────────────────────────────────────────

PAGE_SIZE = 50

@app.get("/orders", response_class=HTMLResponse)
def orders_list(request: Request, status: str = None, search: str = None,
                manager: str = None, factory: str = None,
                date_from: str = None, date_to: str = None,
                page: int = 0,
                db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    q = db.query(Order).filter(Order.deleted_at == None)
    if user.role != "admin":
        q = q.filter(Order.manager_username == user.username)
    if status:
        q = q.filter(Order.status == status)
    if search:
        like = f"%{search}%"
        q = q.filter(
            Order.client_name.ilike(like) | Order.model.ilike(like) |
            Order.factory_name.ilike(like) | Order.client_phone.ilike(like)
        )
    if manager:
        q = q.filter(Order.manager_username == manager)
    if factory:
        q = q.filter(Order.factory_name == factory)
    if date_from:
        try:
            q = q.filter(Order.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import timedelta
            q = q.filter(Order.created_at < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    page = max(0, page)
    total = q.count()
    orders = q.order_by(Order.created_at.desc()).offset(page * PAGE_SIZE).limit(PAGE_SIZE).all()
    counts = {s: db.query(Order).filter(Order.deleted_at == None, Order.status == s).count() for s in STATUS_LABELS}
    managers = [r[0] for r in db.query(Order.manager_username).filter(
        Order.manager_username.isnot(None)).distinct().all()]
    factories = db.query(Factory).order_by(Factory.name).all()
    # base query string без page — для пагинации
    bqs_parts = []
    if status: bqs_parts.append(f"status={status}")
    if search: bqs_parts.append(f"search={search}")
    if manager: bqs_parts.append(f"manager={manager}")
    if factory: bqs_parts.append(f"factory={factory}")
    if date_from: bqs_parts.append(f"date_from={date_from}")
    if date_to: bqs_parts.append(f"date_to={date_to}")
    base_qs = "&".join(bqs_parts)
    return templates.TemplateResponse("orders.html", ctx(request, db, user,
        orders=orders, current_status=status, counts=counts,
        search=search or "", filter_manager=manager or "", filter_factory=factory or "",
        date_from=date_from or "", date_to=date_to or "",
        managers=managers, factories=factories,
        page=page, total=total, page_size=PAGE_SIZE, base_qs=base_qs))


@app.get("/orders/export")
def export_orders(request: Request, status: str = None, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    q = db.query(Order).filter(Order.deleted_at == None)
    if user.role != "admin":
        q = q.filter(Order.manager_username == user.username)
    if status:
        q = q.filter(Order.status == status)
    orders = q.order_by(Order.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = ["#", "Дата", "Договор", "Клиент", "Телефон", "Email", "Фабрика",
               "Категория", "Модель", "Размеры", "Материал", "Цвет",
               "Кол-во", "Срок поставки", "Дата отгрузки", "Способ оплаты",
               "Аванс", "Остаток", "Сумма", "Менеджер", "Статус", "Принято фабрикой"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    status_map = {"new": "Новая", "sent_to_factory": "Отправлено",
                  "accepted": "Принято", "rejected": "Отклонено"}
    for o in orders:
        ws.append([
            o.id,
            o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else "",
            o.contract_number or "",
            o.client_name, o.client_phone or "", o.client_email or "",
            o.factory_name, o.category, o.model,
            o.dimensions or "", o.material or "", o.color or "",
            o.quantity, o.delivery_date or "", o.shipment_date or "",
            o.payment_method or "",
            o.advance_payment or "", o.balance_payment or "",
            o.order_amount or "",
            o.manager_username or "",
            status_map.get(o.status, o.status),
            o.factory_confirmed_at.strftime("%d.%m.%Y") if o.factory_confirmed_at else "",
        ])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    log(db, user, f"Экспорт заявок в Excel ({len(orders)} шт.)", request)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orders.xlsx"})


# ── create order (web form) ───────────────────────────────────────────────────

@app.get("/orders/new", response_class=HTMLResponse)
def order_new_page(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    factories = db.query(Factory).order_by(Factory.name).all()
    categories = ["Диваны", "Кресла", "Кровати", "Шкафы", "Столы", "Стулья", "Тумбы", "Другое"]
    return templates.TemplateResponse("order_form.html", ctx(request, db, user,
        order=None, factories=factories, categories=categories,
        payment_methods=PAYMENT_METHODS))


@app.post("/orders/new")
def order_new_submit(
    request: Request,
    client_name: str = Form(...),
    client_phone: str = Form(""),
    client_phone_name: str = Form(""),
    client_phone2: str = Form(""),
    client_phone2_name: str = Form(""),
    client_whatsapp: str = Form(""),
    client_telegram: str = Form(""),
    client_email: str = Form(""),
    factory_id: int = Form(...),
    category: str = Form(...),
    model: str = Form(...),
    contract_number: str = Form(""),
    contract_date: str = Form(""),
    dimensions: str = Form(""),
    material: str = Form(""),
    color: str = Form(""),
    configuration: str = Form(""),
    quantity: int = Form(1),
    delivery_date: str = Form(""),
    shipment_date: str = Form(""),
    payment_method: str = Form(""),
    advance_payment: str = Form(""),
    balance_payment: str = Form(""),
    delivery_region: str = Form(""),
    delivery_city: str = Form(""),
    delivery_street: str = Form(""),
    delivery_house: str = Form(""),
    delivery_corpus: str = Form(""),
    delivery_apartment: str = Form(""),
    delivery_address_full: str = Form(""),
    comments: str = Form(""),
    files: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    factory = db.query(Factory).filter(Factory.id == factory_id).first()
    if not factory:
        return RedirectResponse("/orders/new?err=Фабрика не найдена", 303)
    adv = _parse_float(advance_payment)
    bal = _parse_float(balance_payment)
    total = (adv or 0) + (bal or 0) or None
    order = Order(
        client_name=client_name.strip(),
        client_phone=client_phone.strip() or None,
        client_phone_name=client_phone_name.strip() or None,
        client_phone2=client_phone2.strip() or None,
        client_phone2_name=client_phone2_name.strip() or None,
        client_whatsapp=client_whatsapp.strip() or None,
        client_telegram=client_telegram.strip() or None,
        client_email=client_email.strip() or None,
        factory_name=factory.name, factory_email=factory.email,
        category=category, model=model.strip(),
        contract_number=contract_number.strip() or None,
        contract_date=contract_date.strip() or None,
        dimensions=dimensions.strip() or None,
        material=material.strip() or None,
        color=color.strip() or None,
        configuration=configuration.strip() or None,
        quantity=quantity,
        delivery_date=delivery_date.strip() or None,
        shipment_date=shipment_date.strip() or None,
        payment_method=payment_method.strip() or None,
        advance_payment=adv, balance_payment=bal, order_amount=total,
        delivery_region=delivery_region.strip() or None,
        delivery_city=delivery_city.strip() or None,
        delivery_street=delivery_street.strip() or None,
        delivery_house=delivery_house.strip() or None,
        delivery_corpus=delivery_corpus.strip() or None,
        delivery_apartment=delivery_apartment.strip() or None,
        delivery_address_full=delivery_address_full.strip() or None,
        manager_id=user.id, manager_username=user.username,
    )
    db.add(order); db.commit(); db.refresh(order)
    log_history(db, order.id, user.username, comment="Заявка создана через веб-форму")
    # Сохраняем вложения
    for f in files:
        if f.filename:
            order_dir = os.path.join(UPLOADS_DIR, str(order.id))
            os.makedirs(order_dir, exist_ok=True)
            ext = os.path.splitext(f.filename)[1]
            saved_name = f"{uuid.uuid4().hex}{ext}"
            with open(os.path.join(order_dir, saved_name), "wb") as out:
                shutil.copyfileobj(f.file, out)
            db.add(OrderAttachment(order_id=order.id, filename=saved_name,
                                   original_name=f.filename, uploaded_by=user.username))
    db.commit()
    log(db, user, f"Создана заявка №{order.id} — {client_name}", request)
    return RedirectResponse(f"/orders/{order.id}?msg=Заявка создана", 303)


@app.get("/orders/{order_id}", response_class=HTMLResponse)
def order_detail(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404)
    attachments = db.query(OrderAttachment).filter(OrderAttachment.order_id == order_id).all()
    commission = round(order.order_amount * COMMISSION_RATE, 2) if order.order_amount else None
    is_owner = user.role == "admin" or order.manager_username == user.username
    return templates.TemplateResponse("order_detail.html", ctx(request, db, user,
        order=order, attachments=attachments, commission=commission,
        commission_rate=int(COMMISSION_RATE * 100), is_owner=is_owner))


@app.post("/orders/{order_id}/confirm")
def confirm_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.new:
        return RedirectResponse(f"/orders/{order_id}?err=Нельзя подтвердить", 303)
    order.confirmation_token = str(uuid.uuid4())
    order.status = OrderStatus.sent_to_factory
    order.sent_to_factory_at = datetime.utcnow()
    log_history(db, order.id, user.username, field="status",
                old_value="new", new_value="sent_to_factory")
    db.commit()
    confirm_url = f"{SITE_URL}/confirm/{order.confirmation_token}"
    try:
        send_factory_email(order, confirm_url)
        msg = "Подтверждено — письмо отправлено фабрике"
    except Exception as e:
        msg = f"Статус обновлён, email не отправлен: {e}"
    log(db, user, f"Подтверждена заявка №{order_id} → фабрика «{order.factory_name}»", request)
    return RedirectResponse(f"/orders/{order_id}?msg={msg}", 303)


@app.post("/orders/{order_id}/reject")
def reject_order(order_id: int, request: Request, reason: str = Form(""),
                 db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status not in [OrderStatus.new, OrderStatus.sent_to_factory]:
        return RedirectResponse(f"/orders/{order_id}?err=Нельзя отклонить", 303)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет прав для изменения чужой заявки", 303)
    old = order.status
    order.status = OrderStatus.rejected
    order.rejection_reason = reason or None
    log_history(db, order.id, user.username, field="status",
                old_value=old, new_value="rejected", comment=reason or None)
    db.commit()
    log(db, user, f"Отклонена заявка №{order_id}", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Заявка отклонена", 303)


@app.post("/orders/{order_id}/upload")
def upload_attachment(order_id: int, request: Request,
                      file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет прав для изменения чужой заявки", 303)
    order_dir = os.path.join(UPLOADS_DIR, str(order_id))
    os.makedirs(order_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "")[1]
    saved_name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(order_dir, saved_name)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    db.add(OrderAttachment(order_id=order_id, filename=saved_name,
                           original_name=file.filename, uploaded_by=user.username))
    db.commit()
    log(db, user, f"Прикреплён файл «{file.filename}» к заявке №{order_id}", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Файл прикреплён", 303)


@app.get("/orders/{order_id}/download/{filename}")
def download_attachment(order_id: int, filename: str, request: Request,
                        db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    att = db.query(OrderAttachment).filter(
        OrderAttachment.order_id == order_id,
        OrderAttachment.filename == filename
    ).first()
    if not att:
        raise HTTPException(404)
    path = os.path.join(UPLOADS_DIR, str(order_id), filename)
    return FileResponse(path, filename=att.original_name or filename)


# ── edit order ────────────────────────────────────────────────────────────────

@app.get("/orders/{order_id}/edit", response_class=HTMLResponse)
def order_edit_page(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет доступа", 303)
    factories = db.query(Factory).order_by(Factory.name).all()
    categories = ["Диваны", "Кресла", "Кровати", "Шкафы", "Столы", "Стулья", "Тумбы", "Другое"]
    return templates.TemplateResponse("order_form.html", ctx(request, db, user,
        order=order, factories=factories, categories=categories,
        payment_methods=PAYMENT_METHODS))


@app.post("/orders/{order_id}/edit")
def order_edit_submit(
    order_id: int, request: Request,
    client_name: str = Form(...),
    client_phone: str = Form(""),
    client_phone_name: str = Form(""),
    client_phone2: str = Form(""),
    client_phone2_name: str = Form(""),
    client_whatsapp: str = Form(""),
    client_telegram: str = Form(""),
    client_email: str = Form(""),
    factory_id: int = Form(...),
    category: str = Form(...),
    model: str = Form(...),
    contract_number: str = Form(""),
    contract_date: str = Form(""),
    dimensions: str = Form(""),
    material: str = Form(""),
    color: str = Form(""),
    configuration: str = Form(""),
    quantity: int = Form(1),
    delivery_date: str = Form(""),
    shipment_date: str = Form(""),
    payment_method: str = Form(""),
    advance_payment: str = Form(""),
    balance_payment: str = Form(""),
    delivery_region: str = Form(""),
    delivery_city: str = Form(""),
    delivery_street: str = Form(""),
    delivery_house: str = Form(""),
    delivery_corpus: str = Form(""),
    delivery_apartment: str = Form(""),
    delivery_address_full: str = Form(""),
    status: str = Form(""),
    rejection_reason: str = Form(""),
    comments: str = Form(""),
    db: Session = Depends(get_db),
):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет доступа", 303)
    factory = db.query(Factory).filter(Factory.id == factory_id).first()
    if not factory:
        return RedirectResponse(f"/orders/{order_id}/edit?err=Фабрика не найдена", 303)

    adv = _parse_float(advance_payment)
    bal = _parse_float(balance_payment)
    total = (adv or 0) + (bal or 0) or None

    # Фиксируем изменения в истории
    changes = {
        "client_name": (order.client_name, client_name.strip()),
        "model": (order.model, model.strip()),
        "advance_payment": (order.advance_payment, adv),
        "balance_payment": (order.balance_payment, bal),
        "order_amount": (order.order_amount, total),
        "payment_method": (order.payment_method, payment_method.strip() or None),
        "shipment_date": (order.shipment_date, shipment_date.strip() or None),
        "contract_number": (order.contract_number, contract_number.strip() or None),
    }
    for field, (old, new) in changes.items():
        if str(old or "") != str(new or ""):
            log_history(db, order.id, user.username, field=field,
                        old_value=old, new_value=new)

    order.client_name = client_name.strip()
    order.client_phone = client_phone.strip() or None
    order.client_phone_name = client_phone_name.strip() or None
    order.client_phone2 = client_phone2.strip() or None
    order.client_phone2_name = client_phone2_name.strip() or None
    order.client_whatsapp = client_whatsapp.strip() or None
    order.client_telegram = client_telegram.strip() or None
    order.client_email = client_email.strip() or None
    order.factory_name = factory.name
    order.factory_email = factory.email
    order.category = category
    order.model = model.strip()
    order.contract_number = contract_number.strip() or None
    order.contract_date = contract_date.strip() or None
    order.dimensions = dimensions.strip() or None
    order.material = material.strip() or None
    order.color = color.strip() or None
    order.configuration = configuration.strip() or None
    order.quantity = quantity
    order.delivery_date = delivery_date.strip() or None
    order.shipment_date = shipment_date.strip() or None
    order.payment_method = payment_method.strip() or None
    order.advance_payment = adv
    order.balance_payment = bal
    order.order_amount = total
    order.comments = comments.strip() or None
    order.delivery_region = delivery_region.strip() or None
    order.delivery_city = delivery_city.strip() or None
    order.delivery_street = delivery_street.strip() or None
    order.delivery_house = delivery_house.strip() or None
    order.delivery_corpus = delivery_corpus.strip() or None
    order.delivery_apartment = delivery_apartment.strip() or None
    order.delivery_address_full = delivery_address_full.strip() or None
    valid_statuses = ["new", "sent_to_factory", "accepted", "shipped", "delivered", "rejected"]
    if status and status in valid_statuses:
        if order.status != status:
            log_history(db, order.id, user.username, field="status",
                        old_value=order.status, new_value=status)
        order.status = status
        order.rejection_reason = rejection_reason.strip() or None if status == "rejected" else order.rejection_reason
    db.commit()
    log(db, user, f"Отредактирована заявка №{order_id}", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Заявка обновлена", 303)


# ── status: shipped / delivered ───────────────────────────────────────────────

@app.post("/orders/{order_id}/ship")
def ship_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.accepted:
        return RedirectResponse(f"/orders/{order_id}?err=Нельзя отгрузить", 303)
    old = order.status
    order.status = OrderStatus.shipped
    log_history(db, order.id, user.username, field="status", old_value=old, new_value="shipped")
    db.commit()
    # Уведомить менеджера
    if order.manager_id:
        mgr = db.query(User).filter(User.id == order.manager_id).first()
        if mgr and mgr.telegram_id:
            send_telegram_message(mgr.telegram_id,
                f"🚚 *Заявка №{order.id} отгружена!*\nКлиент: {order.client_name}\nМодель: {order.model}")
    log(db, user, f"Заявка №{order_id} отмечена как Отгружено", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Статус: Отгружено", 303)


@app.post("/orders/{order_id}/deliver")
def deliver_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order or order.status != OrderStatus.shipped:
        return RedirectResponse(f"/orders/{order_id}?err=Нельзя доставить", 303)
    old = order.status
    order.status = OrderStatus.delivered
    log_history(db, order.id, user.username, field="status", old_value=old, new_value="delivered")
    db.commit()
    if order.manager_id:
        mgr = db.query(User).filter(User.id == order.manager_id).first()
        if mgr and mgr.telegram_id:
            send_telegram_message(mgr.telegram_id,
                f"✅ *Заявка №{order.id} доставлена клиенту!*\nКлиент: {order.client_name}")
    log(db, user, f"Заявка №{order_id} отмечена как Доставлено", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Статус: Доставлено", 303)


# ── PDF генерация ─────────────────────────────────────────────────────────────

@app.get("/orders/{order_id}/pdf")
def order_pdf(order_id: int, request: Request, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(404)
    pdf_bytes = _generate_order_pdf(order)
    log(db, user, f"Сгенерирован PDF для заявки №{order_id}", request)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=order_{order_id}.pdf"}
    )


def _generate_order_pdf(order) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    # Шрифт с поддержкой кириллицы
    font_paths = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    font_name = "Helvetica"
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CustomFont", fp))
                pdfmetrics.registerFont(TTFont("CustomFontB", fp.replace(".ttf", " Bold.ttf") if os.path.exists(fp.replace(".ttf", " Bold.ttf")) else fp))
                font_name = "CustomFont"
                break
            except Exception:
                continue

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontName=font_name, fontSize=16, leading=20,
                                  spaceAfter=6, textColor=colors.HexColor("#0F172A"))
    sub_style = ParagraphStyle("sub", fontName=font_name, fontSize=10,
                                textColor=colors.HexColor("#64748B"), spaceAfter=16)
    label_style = ParagraphStyle("label", fontName=font_name, fontSize=9,
                                  textColor=colors.HexColor("#64748B"))
    value_style = ParagraphStyle("value", fontName=font_name, fontSize=10,
                                  textColor=colors.HexColor("#0F172A"))

    status_map = {"new": "Новая", "sent_to_factory": "Отправлено", "accepted": "Принято",
                  "shipped": "Отгружено", "delivered": "Доставлено", "rejected": "Отклонено"}

    elements = []

    # Заголовок
    elements.append(Paragraph(f"Заявка №{order.id}", title_style))
    date_str = order.created_at.strftime("%d.%m.%Y") if order.created_at else "—"
    elements.append(Paragraph(f"Дата создания: {date_str}  |  Статус: {status_map.get(order.status, order.status)}", sub_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    elements.append(Spacer(1, 0.4*cm))

    def row(label, value):
        return [Paragraph(label, label_style), Paragraph(str(value or "—"), value_style)]

    # Таблица данных
    data = [
        row("Клиент", order.client_name),
        row("Телефон", order.client_phone),
        row("Email клиента", order.client_email),
        row("Фабрика", order.factory_name),
        row("Номер договора", order.contract_number),
        row("Категория", order.category),
        row("Модель / Артикул", order.model),
        row("Размеры", order.dimensions),
        row("Материал / Обивка", order.material),
        row("Цвет", order.color),
        row("Комплектация", order.configuration),
        row("Количество", f"{order.quantity} шт."),
        row("Срок поставки", order.delivery_date),
        row("Дата возможной отгрузки", order.shipment_date),
        row("Способ оплаты", order.payment_method),
    ]
    if order.advance_payment:
        data.append(row("Аванс", f"{order.advance_payment:,.0f} ₽".replace(",", " ")))
    if order.balance_payment:
        data.append(row("Остаток", f"{order.balance_payment:,.0f} ₽".replace(",", " ")))
    if order.order_amount:
        data.append(row("ИТОГО", f"{order.order_amount:,.0f} ₽".replace(",", " ")))
    if order.comments:
        data.append(row("Комментарии", order.comments))
    if order.manager_username:
        data.append(row("Менеджер", order.manager_username))

    t = Table(data, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
        # Выделяем итого
        ("FONTNAME", (0, -1 if order.order_amount else -2), (-1, -1 if order.order_amount else -2), font_name),
        ("TEXTCOLOR", (1, -1 if order.order_amount else -2), (1, -1 if order.order_amount else -2), colors.HexColor("#059669")),
        ("FONTSIZE", (0, -1 if order.order_amount else -2), (-1, -1 if order.order_amount else -2), 11),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 1*cm))

    # Подписи
    sig_data = [["Менеджер: _____________________", "М.П.", "Клиент: _____________________"]]
    sig_t = Table(sig_data, colWidths=[7*cm, 3*cm, 7*cm])
    sig_t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#94A3B8")),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
    ]))
    elements.append(sig_t)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ── История заявки ────────────────────────────────────────────────────────────

@app.get("/orders/{order_id}/history")
def order_history(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    history = db.query(OrderHistory).filter(
        OrderHistory.order_id == order_id
    ).order_by(OrderHistory.created_at.asc()).all()
    return [{"id": h.id, "username": h.username, "field": h.field,
             "old_value": h.old_value, "new_value": h.new_value,
             "comment": h.comment,
             "created_at": h.created_at.strftime("%d.%m.%Y %H:%M") if h.created_at else ""}
            for h in history]


# ── API: поиск по прайсу ──────────────────────────────────────────────────────

@app.get("/api/price-search")
def price_search(q: str = "", db: Session = Depends(get_db)):
    if len(q) < 2:
        return []
    like = f"%{q}%"
    items = db.query(PriceItem).filter(
        PriceItem.name.ilike(like) | PriceItem.article.ilike(like)
    ).limit(20).all()
    return [{"id": i.id, "name": i.name, "article": i.article or "",
             "category": i.category or "", "base_price": i.base_price,
             "markup_price": i.markup_price} for i in items]


# ── Дашборд менеджера ─────────────────────────────────────────────────────────

@app.get("/my-stats", response_class=HTMLResponse)
def my_stats(request: Request, db: Session = Depends(get_db)):
    import calendar
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)

    all_orders = db.query(Order).filter(Order.deleted_at == None, Order.manager_username == user.username).all()
    month_orders = [o for o in all_orders
                    if o.factory_confirmed_at and o.factory_confirmed_at >= month_start
                    and o.status in ("accepted", "shipped", "delivered")]
    total_orders = [o for o in all_orders if o.status in ("accepted", "shipped", "delivered")]

    month_revenue = sum(o.order_amount or 0 for o in month_orders)
    total_revenue = sum(o.order_amount or 0 for o in total_orders)

    month_names = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
    monthly_data = []
    for i in range(5, -1, -1):
        m = (now.month - 1 - i) % 12 + 1
        y = now.year + ((now.month - 1 - i) // 12)
        first = datetime(y, m, 1)
        last_day = calendar.monthrange(y, m)[1]
        last = datetime(y, m, last_day, 23, 59, 59)
        mo = [o for o in all_orders
              if o.factory_confirmed_at and first <= o.factory_confirmed_at <= last
              and o.status in ("accepted", "shipped", "delivered")]
        total = sum(o.order_amount or 0 for o in mo)
        monthly_data.append({
            "label": f"{month_names[m-1]} {str(y)[2:]}",
            "total": round(total),
            "commission": round(total * COMMISSION_RATE),
            "count": len(mo),
        })

    counts = {s: sum(1 for o in all_orders if o.status == s) for s in STATUS_LABELS}
    recent = sorted(all_orders, key=lambda o: o.created_at or datetime.min, reverse=True)[:10]

    return templates.TemplateResponse("my_stats.html", ctx(request, db, user,
        month_revenue=month_revenue, total_revenue=total_revenue,
        month_commission=round(month_revenue * COMMISSION_RATE, 2),
        total_commission=round(total_revenue * COMMISSION_RATE, 2),
        month_count=len(month_orders), total_count=len(total_orders),
        monthly_data=monthly_data, counts=counts, recent=recent,
        commission_rate=int(COMMISSION_RATE * 100)))


# ── factory confirmation (public) ─────────────────────────────────────────────

@app.get("/confirm/{token}", response_class=HTMLResponse)
def factory_confirm(token: str, request: Request, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.confirmation_token == token).first()
    if not order:
        return templates.TemplateResponse("confirm_result.html",
            {"request": request, "success": False, "message": "Ссылка недействительна."})
    if order.status == OrderStatus.accepted:
        return templates.TemplateResponse("confirm_result.html",
            {"request": request, "success": True, "order": order,
             "message": f"Заявка №{order.id} уже подтверждена."})
    order.status = OrderStatus.accepted
    order.factory_confirmed_at = datetime.utcnow()
    log_history(db, order.id, "factory", field="status",
                old_value="sent_to_factory", new_value="accepted")
    db.commit()
    try:
        send_admin_notification(order)
    except Exception:
        pass
    # Уведомить менеджера в Telegram если у него есть chat_id
    if order.manager_id:
        manager = db.query(User).filter(User.id == order.manager_id).first()
        if manager and manager.telegram_id:
            amount_str = f"{int(order.order_amount):,} ₽".replace(",", " ") if order.order_amount else "—"
            send_telegram_message(
                manager.telegram_id,
                f"✅ *Заявка №{order.id} подтверждена фабрикой!*\n\n"
                f"Клиент: {order.client_name}\n"
                f"Фабрика: {order.factory_name}\n"
                f"Модель: {order.model}\n"
                f"Сумма: {amount_str}"
            )
    return templates.TemplateResponse("confirm_result.html",
        {"request": request, "success": True, "order": order,
         "message": f"Заявка №{order.id} подтверждена. Спасибо!"})


# ── price list ────────────────────────────────────────────────────────────────

@app.get("/pricelist", response_class=HTMLResponse)
def pricelist_page(request: Request, batch_id: str = None, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    batches = db.query(PriceBatch).order_by(PriceBatch.uploaded_at.desc()).all()
    items = []
    selected = None
    if batch_id:
        selected = db.query(PriceBatch).filter(PriceBatch.uuid == batch_id).first()
        if selected:
            items = db.query(PriceItem).filter(PriceItem.batch_uuid == batch_id).all()
    elif batches:
        selected = batches[0]
        items = db.query(PriceItem).filter(PriceItem.batch_uuid == selected.uuid).all()
    return templates.TemplateResponse("pricelist.html", ctx(request, db, user,
        batches=batches, items=items, selected=selected))


@app.post("/pricelist/upload")
async def upload_pricelist(request: Request, factory_name: str = Form(""),
                           file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/pricelist", 303)
    contents = await file.read()
    try:
        items = parse_price_file(contents, file.filename or "")
    except Exception as e:
        return RedirectResponse(f"/pricelist?err={e}", 303)
    if not items:
        return RedirectResponse("/pricelist?err=Файл пустой или не распознан", 303)
    batch_uuid = str(uuid.uuid4())
    batch = PriceBatch(uuid=batch_uuid, factory_name=factory_name or file.filename,
                       filename=file.filename, item_count=len(items),
                       uploaded_by=user.username)
    db.add(batch)
    db.bulk_insert_mappings(PriceItem, [
        {"batch_uuid": batch_uuid, **item} for item in items
    ])
    db.commit()
    log(db, user, f"Загружен прайс «{file.filename}» ({len(items)} позиций)", request)
    return RedirectResponse(f"/pricelist?batch_id={batch_uuid}&msg=Прайс загружен: {len(items)} позиций", 303)


@app.post("/pricelist/{batch_uuid}/delete")
def delete_pricelist(batch_uuid: str, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/pricelist", 303)
    batch = db.query(PriceBatch).filter(PriceBatch.uuid == batch_uuid).first()
    if batch:
        db.query(PriceItem).filter(PriceItem.batch_uuid == batch_uuid).delete()
        db.delete(batch)
        db.commit()
        log(db, user, f"Удалён прайс «{batch.filename}»", request)
    return RedirectResponse("/pricelist?msg=Прайс удалён", 303)


# ── report ────────────────────────────────────────────────────────────────────

@app.get("/report", response_class=HTMLResponse)
def report_page(request: Request, month: int = None, year: int = None,
                db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)

    now = datetime.utcnow()
    month = month or now.month
    year  = year  or now.year

    q = db.query(Order).filter(
        Order.status == OrderStatus.accepted,
        Order.order_amount.isnot(None),
    )
    # Admin sees all; manager sees own
    if user.role != "admin":
        q = q.filter(Order.manager_id == user.id)

    # Filter by month/year using Python (SQLite has limited date funcs)
    all_orders = q.all()
    period_orders = [
        o for o in all_orders
        if o.factory_confirmed_at
        and o.factory_confirmed_at.month == month
        and o.factory_confirmed_at.year == year
    ]

    # Group by manager
    managers = {}
    for o in period_orders:
        key = o.manager_username or "—"
        if key not in managers:
            managers[key] = {"username": key, "orders": [], "total": 0, "commission": 0}
        managers[key]["orders"].append(o)
        managers[key]["total"] += o.order_amount or 0
        managers[key]["commission"] += (o.order_amount or 0) * COMMISSION_RATE

    rows = sorted(managers.values(), key=lambda x: x["commission"], reverse=True)
    for r in rows:
        r["commission"] = round(r["commission"], 2)
        r["total"] = round(r["total"], 2)

    # Month navigation list
    months = ["Январь","Февраль","Март","Апрель","Май","Июнь",
              "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    return templates.TemplateResponse("report.html", ctx(request, db, user,
        rows=rows, month=month, year=year, months=months,
        total_amount=sum(r["total"] for r in rows),
        total_commission=round(sum(r["commission"] for r in rows), 2),
        commission_rate=int(COMMISSION_RATE * 100),
        period_orders=period_orders))


# ── dashboard (admin) ─────────────────────────────────────────────────────────

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    from datetime import timedelta
    import calendar
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    counts = {s: db.query(Order).filter(Order.deleted_at == None, Order.status == s).count() for s in STATUS_LABELS}
    total_revenue = sum(
        o.order_amount or 0 for o in db.query(Order).filter(Order.deleted_at == None, Order.status == "accepted").all()
    )
    # Динамика за 6 месяцев
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    month_revenue = sum(
        o.order_amount or 0 for o in db.query(Order).filter(
            Order.deleted_at == None,
            Order.status == "accepted",
            Order.factory_confirmed_at >= month_start,
        ).all()
    )
    monthly_data = []
    month_names = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
    for i in range(5, -1, -1):
        month = (now.month - 1 - i) % 12 + 1
        year = now.year + ((now.month - 1 - i) // 12)
        first = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        last = datetime(year, month, last_day, 23, 59, 59)
        month_orders = db.query(Order).filter(
            Order.deleted_at == None,
            Order.factory_confirmed_at >= first,
            Order.factory_confirmed_at <= last,
            Order.status == "accepted"
        ).all()
        total = sum(o.order_amount or 0 for o in month_orders)
        monthly_data.append({
            "label": f"{month_names[month-1]} {str(year)[2:]}",
            "total": round(total),
            "commission": round(total * COMMISSION_RATE),
            "count": len(month_orders),
        })
    # Топ менеджеры
    mgr_rows = []
    for (uname,) in db.query(Order.manager_username).filter(
            Order.deleted_at == None, Order.manager_username.isnot(None), Order.status == "accepted").distinct().all():
        mgr_orders = db.query(Order).filter(
            Order.deleted_at == None, Order.manager_username == uname, Order.status == "accepted").all()
        total = sum(o.order_amount or 0 for o in mgr_orders)
        mgr_rows.append({"username": uname, "orders": len(mgr_orders),
                         "total": round(total), "commission": round(total * COMMISSION_RATE)})
    mgr_rows.sort(key=lambda x: x["total"], reverse=True)
    # Ожидают ответа > 3 дней
    overdue = db.query(Order).filter(
        Order.deleted_at == None,
        Order.status == "sent_to_factory",
        Order.sent_to_factory_at <= datetime.utcnow() - timedelta(days=3)
    ).all()
    return templates.TemplateResponse("dashboard.html", ctx(request, db, user,
        counts=counts, total_revenue=total_revenue, month_revenue=month_revenue,
        monthly_data=monthly_data, mgr_rows=mgr_rows, overdue=overdue,
        commission_rate=int(COMMISSION_RATE * 100)))


@app.post("/admin/send-reminders")
def send_reminders(request: Request, db: Session = Depends(get_db)):
    from datetime import timedelta
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    overdue = db.query(Order).filter(
        Order.status == "sent_to_factory",
        Order.sent_to_factory_at <= datetime.utcnow() - timedelta(days=3)
    ).all()
    sent = 0
    for order in overdue:
        try:
            confirm_url = f"{SITE_URL}/confirm/{order.confirmation_token}"
            send_factory_email(order, confirm_url)
            sent += 1
        except Exception:
            pass
    log(db, user, f"Напоминания фабрикам: {sent} писем", request)
    return RedirectResponse(f"/dashboard?msg=Отправлено напоминаний: {sent}", 303)


# ── factories (admin) ─────────────────────────────────────────────────────────

@app.get("/factories", response_class=HTMLResponse)
def factories_list(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    factories = db.query(Factory).order_by(Factory.name).all()
    return templates.TemplateResponse("factories.html", ctx(request, db, user, factories=factories))


@app.post("/factories/add")
def add_factory(request: Request, name: str = Form(...), email: str = Form(...),
                reply_email: str = Form(""), db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    db.add(Factory(name=name.strip(), email=email.strip(),
                   reply_email=reply_email.strip() or None))
    db.commit()
    log(db, user, f"Добавлена фабрика «{name}»", request)
    return RedirectResponse("/factories?msg=Фабрика добавлена", 303)


@app.post("/factories/{factory_id}/delete")
def delete_factory(factory_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    f = db.query(Factory).filter(Factory.id == factory_id).first()
    if f:
        log(db, user, f"Удалена фабрика «{f.name}»", request)
        db.delete(f); db.commit()
    return RedirectResponse("/factories?msg=Фабрика удалена", 303)


# ── users (admin) ─────────────────────────────────────────────────────────────

@app.get("/users", response_class=HTMLResponse)
def users_list(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    users = db.query(User).order_by(User.created_at).all()
    return templates.TemplateResponse("users.html", ctx(request, db, user, users=users))


@app.post("/users/add")
def add_user(request: Request, username: str = Form(...), display_name: str = Form(""),
             password: str = Form(...), role: str = Form("manager"),
             db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    if db.query(User).filter(User.username == username).first():
        users = db.query(User).order_by(User.created_at).all()
        return templates.TemplateResponse("users.html", ctx(request, db, user,
            users=users, error=f"Пользователь «{username}» уже существует"))
    db.add(User(username=username.strip(), password_hash=hash_password(password),
                display_name=display_name.strip() or None, role=role))
    db.commit()
    log(db, user, f"Создан пользователь «{username}» (роль: {role})", request)
    return RedirectResponse("/users?msg=Пользователь создан", 303)


@app.post("/users/{user_id}/reset-password")
def reset_password(user_id: int, request: Request, new_password: str = Form(...),
                   db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    target = db.query(User).filter(User.id == user_id).first()
    if target:
        target.password_hash = hash_password(new_password)
        db.commit()
        log(db, user, f"Сброшен пароль «{target.username}»", request)
    return RedirectResponse("/users?msg=Пароль обновлён", 303)


@app.post("/users/{user_id}/toggle")
def toggle_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    target = db.query(User).filter(User.id == user_id).first()
    if target and target.role != "admin":
        target.is_active = not target.is_active
        db.commit()
        log(db, user, f"{'Активирован' if target.is_active else 'Деактивирован'} «{target.username}»", request)
    return RedirectResponse("/users?msg=Статус обновлён", 303)


@app.post("/users/{user_id}/delete")
def delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    target = db.query(User).filter(User.id == user_id).first()
    if target and target.role != "admin":
        log(db, user, f"Удалён «{target.username}»", request)
        db.delete(target); db.commit()
    return RedirectResponse("/users?msg=Пользователь удалён", 303)


@app.post("/users/me/update")
def update_my_profile(request: Request,
                      display_name: str = Form(""),
                      new_username: str = Form(""),
                      new_password: str = Form(""),
                      db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    if display_name.strip():
        user.display_name = display_name.strip()
    if new_username.strip() and new_username.strip() != user.username:
        if db.query(User).filter(User.username == new_username.strip()).first():
            users = db.query(User).order_by(User.created_at).all()
            return templates.TemplateResponse("users.html", ctx(request, db, user,
                users=users, profile_error="Логин уже занят"))
        user.username = new_username.strip()
    if new_password.strip():
        user.password_hash = hash_password(new_password.strip())
    db.commit()
    log(db, user, "Обновил свой профиль", request)
    # Обновляем сессию с новым логином
    from fastapi.responses import Response
    from app.auth import make_token
    resp = RedirectResponse("/users?msg=Профиль обновлён", 303)
    token = make_token(user.id, user.username, user.role)
    resp.set_cookie("session", token, httponly=True, max_age=86400*30, samesite="lax")
    return resp


# ── activity (admin) ──────────────────────────────────────────────────────────

@app.get("/activity", response_class=HTMLResponse)
def activity_log(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    logs = db.query(ActivityLog).order_by(ActivityLog.created_at.desc()).limit(500).all()
    return templates.TemplateResponse("activity.html", ctx(request, db, user, logs=logs))


# ── settings (admin) ──────────────────────────────────────────────────────────

@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    s = get_settings(db)
    return templates.TemplateResponse("settings.html", ctx(request, db, user, settings=s))


@app.post("/settings")
def save_settings(request: Request, theme: str = Form("light"),
                  primary_color: str = Form("indigo"), db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    s = get_settings(db)
    s.theme = theme; s.primary_color = primary_color
    db.commit()
    log(db, user, f"Настройки: тема={theme}, цвет={primary_color}", request)
    return RedirectResponse("/settings?msg=Настройки сохранены", 303)


# ── API для Telegram-бота ─────────────────────────────────────────────────────

@app.post("/api/verify-manager")
async def verify_manager(request: Request, db: Session = Depends(get_db)):
    data = await request.json()
    user = db.query(User).filter(User.username == data.get("username")).first()
    if not user or not user.is_active or not verify_password(data.get("password", ""), user.password_hash):
        return {"ok": False, "error": "Неверный логин или пароль"}
    if user.role not in ("admin", "manager"):
        return {"ok": False, "error": "Нет доступа"}
    # Сохраняем Telegram chat_id менеджера для уведомлений
    telegram_id = str(data.get("telegram_id", "")).strip()
    if telegram_id and user.telegram_id != telegram_id:
        user.telegram_id = telegram_id
        db.commit()
    return {"ok": True, "user_id": user.id, "username": user.username,
            "display_name": user.display_name or user.username, "role": user.role}


@app.post("/api/orders")
async def create_order_api(request: Request, db: Session = Depends(get_db)):
    data = await request.json()
    factory = db.query(Factory).filter(Factory.name == data.get("factory_name")).first()
    if not factory:
        return {"ok": False, "error": "Фабрика не найдена"}
    advance = data.get("advance_payment") or 0
    balance = data.get("balance_payment") or 0
    total = advance + balance if (advance or balance) else data.get("order_amount")
    order = Order(
        client_name=data["client_name"],
        client_phone=data.get("client_phone"),
        client_email=data.get("client_email"),
        client_telegram_id=str(data.get("client_telegram_id", "")),
        factory_name=factory.name, factory_email=factory.email,
        category=data["category"], model=data["model"],
        dimensions=data.get("dimensions"), material=data.get("material"),
        color=data.get("color"), configuration=data.get("configuration"),
        quantity=data.get("quantity", 1), delivery_date=data.get("delivery_date"),
        comments=data.get("comments"), photo_url=data.get("photo_url"),
        contract_number=data.get("contract_number"),
        shipment_date=data.get("shipment_date"),
        payment_method=data.get("payment_method"),
        advance_payment=advance or None,
        balance_payment=balance or None,
        order_amount=total,
        manager_id=data.get("manager_id"),
        manager_username=data.get("manager_username"),
    )
    db.add(order); db.commit(); db.refresh(order)
    return {"ok": True, "id": order.id}


@app.get("/api/factories")
def api_factories(db: Session = Depends(get_db)):
    return [{"id": f.id, "name": f.name} for f in db.query(Factory).order_by(Factory.name).all()]


# ── Удаление заявки (в корзину) ───────────────────────────────────────────────

@app.post("/orders/{order_id}/delete")
def delete_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at == None).first()
    if not order:
        return RedirectResponse("/orders?err=Заявка не найдена", 303)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse("/orders?err=Нет прав", 303)
    order.deleted_at = datetime.utcnow()
    order.deleted_by = user.username
    log_history(db, order.id, user.username, comment="Заявка перемещена в корзину")
    db.commit()
    log(db, user, f"Заявка №{order_id} перемещена в корзину", request)
    return RedirectResponse("/orders?msg=Заявка перемещена в корзину", 303)


# ── Удаление вложения ─────────────────────────────────────────────────────────

@app.post("/orders/{order_id}/attachments/{att_id}/delete")
def delete_attachment(order_id: int, att_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if order and user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет прав для изменения чужой заявки", 303)
    att = db.query(OrderAttachment).filter(
        OrderAttachment.id == att_id,
        OrderAttachment.order_id == order_id
    ).first()
    if att:
        path = os.path.join(UPLOADS_DIR, str(order_id), att.filename)
        if os.path.exists(path):
            os.remove(path)
        db.delete(att)
        db.commit()
        log(db, user, f"Удалено вложение «{att.original_name}» из заявки №{order_id}", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Файл удалён", 303)


# ── Корзина ───────────────────────────────────────────────────────────────────

@app.get("/trash", response_class=HTMLResponse)
def trash_page(request: Request, db: Session = Depends(get_db)):
    from datetime import timedelta
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    cutoff = datetime.utcnow() - timedelta(days=90)
    q = db.query(Order).filter(Order.deleted_at != None, Order.deleted_at >= cutoff)
    if user.role != "admin":
        q = q.filter(Order.manager_username == user.username)
    orders = q.order_by(Order.deleted_at.desc()).all()
    trash_count = db.query(Order).filter(Order.deleted_at != None, Order.deleted_at < cutoff).count()
    return templates.TemplateResponse("trash.html", ctx(request, db, user,
        orders=orders, expired_count=trash_count))


@app.post("/trash/{order_id}/restore")
def restore_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id, Order.deleted_at != None).first()
    if not order:
        return RedirectResponse("/trash?err=Заявка не найдена", 303)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse("/trash?err=Нет прав", 303)
    order.deleted_at = None
    order.deleted_by = None
    log_history(db, order.id, user.username, comment="Заявка восстановлена из корзины")
    db.commit()
    log(db, user, f"Заявка №{order_id} восстановлена из корзины", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Заявка восстановлена", 303)


@app.post("/trash/empty")
def empty_trash(request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/trash", 303)
    deleted_orders = db.query(Order).filter(Order.deleted_at != None).all()
    count = 0
    for order in deleted_orders:
        order_dir = os.path.join(UPLOADS_DIR, str(order.id))
        if os.path.exists(order_dir):
            shutil.rmtree(order_dir)
        db.query(OrderAttachment).filter(OrderAttachment.order_id == order.id).delete()
        db.delete(order)
        count += 1
    db.commit()
    log(db, user, f"Корзина очищена: удалено {count} заявок", request)
    return RedirectResponse(f"/trash?msg=Корзина очищена ({count} заявок)", 303)


# ── Отмена заявки ─────────────────────────────────────────────────────────────

@app.post("/orders/{order_id}/cancel")
def cancel_order(order_id: int, request: Request,
                 reason: str = Form(""), db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        return RedirectResponse("/orders", 303)
    if user.role != "admin" and order.manager_username != user.username:
        return RedirectResponse(f"/orders/{order_id}?err=Нет прав для изменения чужой заявки", 303)
    old = order.status
    order.status = "cancelled"
    order.rejection_reason = reason.strip() or None
    log_history(db, order.id, user.username, field="status", old_value=old, new_value="cancelled",
                comment=reason.strip() or "Заявка отменена")
    db.commit()
    log(db, user, f"Заявка №{order_id} отменена", request)
    return RedirectResponse(f"/orders/{order_id}?msg=Заявка отменена", 303)


# ── Дублирование заявки ───────────────────────────────────────────────────────

@app.post("/orders/{order_id}/duplicate")
def duplicate_order(order_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_user(request, db)
    if not user:
        return RedirectResponse("/login", 303)
    src = db.query(Order).filter(Order.id == order_id).first()
    if not src:
        return RedirectResponse("/orders", 303)
    new_order = Order(
        client_name=src.client_name, client_phone=src.client_phone,
        client_phone_name=src.client_phone_name, client_phone2=src.client_phone2,
        client_phone2_name=src.client_phone2_name, client_whatsapp=src.client_whatsapp,
        client_telegram=src.client_telegram, client_email=src.client_email,
        factory_name=src.factory_name, factory_email=src.factory_email,
        category=src.category, model=src.model, dimensions=src.dimensions,
        material=src.material, color=src.color, configuration=src.configuration,
        quantity=src.quantity, delivery_date=src.delivery_date, comments=src.comments,
        contract_number=src.contract_number, shipment_date=src.shipment_date,
        payment_method=src.payment_method, advance_payment=src.advance_payment,
        balance_payment=src.balance_payment, order_amount=src.order_amount,
        delivery_region=src.delivery_region, delivery_city=src.delivery_city,
        delivery_street=src.delivery_street, delivery_house=src.delivery_house,
        delivery_corpus=src.delivery_corpus, delivery_apartment=src.delivery_apartment,
        delivery_address_full=src.delivery_address_full,
        manager_id=user.id, manager_username=user.username,
        status="new",
    )
    db.add(new_order); db.commit(); db.refresh(new_order)
    log_history(db, new_order.id, user.username, comment=f"Создана как копия заявки №{order_id}")
    db.commit()
    log(db, user, f"Заявка №{new_order.id} создана как копия №{order_id}", request)
    return RedirectResponse(f"/orders/{new_order.id}/edit?msg=Копия заявки создана", 303)


# ── Напоминания о сроках ──────────────────────────────────────────────────────

@app.get("/admin/reminders", response_class=HTMLResponse)
def send_reminders(request: Request, db: Session = Depends(get_db)):
    from datetime import timedelta
    user = get_user(request, db)
    if not user or user.role != "admin":
        return RedirectResponse("/orders", 303)
    today = datetime.utcnow().date()
    threshold = str(today + timedelta(days=3))
    orders = db.query(Order).filter(
        Order.status.in_(["new", "sent_to_factory", "accepted"]),
        Order.delivery_date <= threshold,
        Order.delivery_date >= str(today)
    ).all()
    sent = 0
    for order in orders:
        if order.manager_id:
            mgr = db.query(User).filter(User.id == order.manager_id).first()
            if mgr and mgr.telegram_id:
                send_telegram_message(mgr.telegram_id,
                    f"⏰ *Напоминание: срок поставки*\n"
                    f"Заявка №{order.id} — {order.client_name}\n"
                    f"Модель: {order.model}\n"
                    f"Срок: {order.delivery_date}\n"
                    f"Статус: {STATUS_LABELS.get(order.status, order.status)}")
                sent += 1
    log(db, user, f"Отправлено {sent} напоминаний о сроках", request)
    return RedirectResponse(f"/orders?msg=Отправлено напоминаний: {sent}", 303)
