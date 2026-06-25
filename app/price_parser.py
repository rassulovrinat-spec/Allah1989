"""
Профессиональный парсер прайс-листов: Excel (.xlsx/.xls) и PDF.
- Автоматически находит строку заголовков
- Поддерживает мёрженные ячейки Excel (категории из подзаголовков)
- Нечёткий поиск колонок по ключевым словам
- Дедупликация, фильтрация мусора
- Любые числовые форматы цен
"""
import re
from io import BytesIO


# ── Публичный API ──────────────────────────────────────────────────────────────

def parse_price_file(file_bytes: bytes, filename: str = "") -> list:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "xlsx"
    if ext == "pdf":
        return _parse_pdf(file_bytes)
    return _parse_excel(file_bytes)


# ── Excel ──────────────────────────────────────────────────────────────────────

def _parse_excel(file_bytes: bytes) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    all_items = []
    for ws in wb.worksheets:
        # Собираем мёрженные диапазоны для восстановления значений
        merged_map = _build_merged_map(ws)
        rows = []
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                val = cell.value
                if val is None and (cell.row, cell.column) in merged_map:
                    val = merged_map[(cell.row, cell.column)]
                cells.append(val)
            rows.append(tuple(cells))
        if not rows:
            continue
        items = _extract_from_rows(rows)
        all_items.extend(items)

    if not all_items:
        raise ValueError(
            "Не удалось распознать прайс-лист. "
            "Убедитесь, что файл содержит колонку с ценами (число > 0)."
        )
    return _deduplicate(all_items)


def _build_merged_map(ws) -> dict:
    """Возвращает {(row, col): value} для всех мёрженных ячеек."""
    result = {}
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r != rng.min_row or c != rng.min_col:
                    result[(r, c)] = top_left
    return result


# ── PDF ────────────────────────────────────────────────────────────────────────

def _parse_pdf(file_bytes: bytes) -> list:
    import pdfplumber
    all_items = []

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            # Сначала пробуем таблицы с разными стратегиями
            for strategy in [None, {"vertical_strategy": "text", "horizontal_strategy": "text"}]:
                try:
                    tables = page.extract_tables(strategy) if strategy else page.extract_tables()
                    for table in (tables or []):
                        if not table or len(table) < 2:
                            continue
                        rows = [tuple(c if c is not None else "" for c in r) for r in table]
                        items = _extract_from_rows(rows)
                        all_items.extend(items)
                    if all_items:
                        break
                except Exception:
                    continue

        # Fallback — текстовые строки
        if not all_items:
            with pdfplumber.open(BytesIO(file_bytes)) as pdf2:
                for page in pdf2.pages:
                    text_rows = _pdf_text_to_rows(page)
                    if text_rows:
                        items = _extract_from_rows(text_rows)
                        all_items.extend(items)

    result = _deduplicate(all_items)
    if not result:
        raise ValueError(
            "Не удалось извлечь данные из PDF. "
            "Убедитесь, что PDF содержит таблицы с ценами."
        )
    return result


def _pdf_text_to_rows(page) -> list:
    text = page.extract_text() or ""
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    rows = []
    for line in lines:
        parts = re.split(r"\s{2,}|\t", line)
        if len(parts) >= 2:
            rows.append(tuple(parts))
    return rows


# ── Дедупликация ───────────────────────────────────────────────────────────────

def _deduplicate(items: list) -> list:
    seen = set()
    result = []
    for item in items:
        key = (item.get("article"), item["name"], item["base_price"])
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


# ── Универсальная логика поиска данных ────────────────────────────────────────

def _extract_from_rows(rows: list) -> list:
    if not rows:
        return []

    header_idx = _find_header_row(rows)
    if header_idx is None:
        header_idx = 0

    headers = [_norm(c) for c in rows[header_idx]]

    price_col    = _find_col(headers, PRICE_KW)
    article_col  = _find_col(headers, ARTICLE_KW)
    name_col     = _find_col(headers, NAME_KW)
    category_col = _find_col(headers, CATEGORY_KW)
    dims_col     = _find_col(headers, DIMS_KW)

    # Если name не найден — берём первый нечисловой столбец после артикула
    if name_col is None:
        for i, h in enumerate(headers):
            if i == article_col:
                continue
            if h and not _as_price(h):
                name_col = i
                break

    if price_col is None:
        price_col = _detect_price_col(rows, header_idx + 1)

    if price_col is None:
        return []

    items = []
    current_category = None  # категория из подзаголовков

    for row in rows[header_idx + 1:]:
        if not any(_norm(c) for c in row):
            continue

        raw_price = _cell(row, price_col)
        base = _as_price(raw_price)

        # Строка-подзаголовок (категория): нет цены, одна непустая ячейка
        if base is None or base <= 0:
            non_empty = [_cell(row, i) for i in range(len(row)) if _cell(row, i)]
            if len(non_empty) == 1 and len(non_empty[0]) > 2:
                current_category = non_empty[0]
            elif len(non_empty) >= 1 and category_col is None:
                # Возможно это категория в первой колонке
                first = _cell(row, 0)
                if first and not _as_price(first) and len(first) > 2:
                    current_category = first
            continue

        name = _cell(row, name_col)
        if not name or _as_price(name) is not None:
            continue

        if _is_junk_row(name, base):
            continue

        article  = _cell(row, article_col)
        category = _cell(row, category_col) or current_category or ""
        dims     = _cell(row, dims_col)

        full_name = f"{name} ({dims})" if dims and dims not in name else name

        items.append({
            "article":      article or None,
            "name":         full_name.strip(),
            "category":     category.strip() or None,
            "base_price":   base,
            "markup_price": round(base * 1.3, 2),
        })

    return items


def _find_header_row(rows: list):
    """Ищет строку с заголовками — максимум совпадений с ключевыми словами."""
    all_kw = set(PRICE_KW + NAME_KW + ARTICLE_KW + CATEGORY_KW + DIMS_KW)
    best_idx, best_score = None, 0
    for i, row in enumerate(rows[:30]):  # заголовок не дальше 30-й строки
        cells = [_norm(c) for c in row]
        score = sum(1 for c in cells if any(kw in c for kw in all_kw))
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx if best_score >= 1 else None


def _detect_price_col(rows: list, start: int) -> int:
    """Определяет колонку с ценами по содержимому данных."""
    col_prices = {}
    for row in rows[start:start + 30]:
        for i, cell in enumerate(row):
            v = _as_price(str(cell or ""))
            if v and 10 < v < 100_000_000:  # реалистичный диапазон цен
                col_prices[i] = col_prices.get(i, 0) + 1
    if not col_prices:
        return None
    return max(col_prices, key=col_prices.get)


def _is_junk_row(name: str, price: float) -> bool:
    """Фильтрует строки-мусор: итоги, заголовки, пустышки."""
    if len(name.strip()) < 2:
        return True
    junk_patterns = [
        r"^итого", r"^всего", r"^total", r"^sum",
        r"^№$", r"^n$", r"^п/п$",
    ]
    n = name.strip().lower()
    return any(re.match(p, n) for p in junk_patterns)


# ── Словари ключевых слов ─────────────────────────────────────────────────────

PRICE_KW    = ["цена", "price", "стоимость", "прайс", "оптовая", "розн", "опт",
               "руб", "rub", "₽", "opt", "розница", "оптом"]
ARTICLE_KW  = ["артикул", "арт.", "арт ", "код", "article", "sku", "id", "№", "art"]
NAME_KW     = ["наименование", "название", "товар", "модель", "продукт",
               "name", "product", "item", "description", "номенклатура"]
CATEGORY_KW = ["категория", "раздел", "группа", "тип", "вид",
               "category", "type", "section", "подраздел"]
DIMS_KW     = ["размер", "габарит", "dimension", "ш×", "ш x", "ш*",
               "дхш", "дхв", "длина", "ширина", "высота", "wxh", "lxw", "мм", "см"]


# ── Вспомогательные функции ───────────────────────────────────────────────────

def _norm(val) -> str:
    if val is None:
        return ""
    return str(val).strip().lower()


def _cell(row, col) -> str:
    if col is None or col >= len(row):
        return ""
    v = row[col]
    return str(v).strip() if v is not None else ""


def _find_col(headers: list, keywords: list):
    """Находит индекс колонки — точное совпадение, затем вхождение."""
    # Точное совпадение
    for i, h in enumerate(headers):
        if h in keywords:
            return i
    # Вхождение подстроки
    for i, h in enumerate(headers):
        for kw in keywords:
            if kw in h:
                return i
    return None


_PRICE_CLEAN = re.compile(r"[^\d.,]")
_PRICE_RANGE = re.compile(r"(\d[\d\s.,]+)\s*[-–—]\s*\d")


def _as_price(val: str):
    """
    Парсит цену из строки любого формата:
    45000 / 45 000 / 45.000 / 45,000.00 / 45 000 руб.
    """
    if not val:
        return None
    s = str(val).strip()

    # Диапазон — берём первое значение
    m = _PRICE_RANGE.search(s)
    if m:
        s = m.group(1)

    s = _PRICE_CLEAN.sub("", s).replace(" ", "")
    if not s:
        return None

    if "," in s and "." in s:
        if s.index(",") > s.index("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")

    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


# Обратная совместимость
def parse_excel(file_bytes: bytes) -> list:
    return _parse_excel(file_bytes)
